from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from django.utils import timezone
from django.db.models import Q
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pandas as pd
from io import BytesIO
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.contrib.auth import get_user_model, login
from rest_framework.permissions import AllowAny
import logging
from rest_framework.pagination import PageNumberPagination
from django.db.models import Prefetch
from django.db import transaction
from datetime import datetime
from core.permissions import IsAdminUser, IsAuthenticatedUser, IsOwnerOrAdmin

from .models import Equipment, Rental, RentalRequest, EquipmentMacAddress, EquipmentHistory
from .serializers import EquipmentSerializer, RentalSerializer, RentalRequestSerializer, EquipmentMacAddressSerializer, EquipmentLiteSerializer, EquipmentHistorySerializer
from django.contrib.auth import get_user_model
from devices.models import Device

User = get_user_model()

class IsAdminOrReadOnly(permissions.BasePermission):
    """
    관리자만 쓰기 권한을 가지며, 읽기는 인증된 사용자에게 허용
    """
    def has_permission(self, request, view):
        # register, by_mac 액션은 인증 없이 허용
        if getattr(view, 'action', None) in ['register', 'by_mac']:
            return True
            
        if request.method in permissions.SAFE_METHODS:
            return request.user.is_authenticated
        return request.user.is_authenticated and request.user.is_superuser


class IsOwnerOrAdmin(permissions.BasePermission):
    """
    본인의 자원이거나 관리자인 경우 접근 가능
    """
    def has_object_permission(self, request, view, obj):
        if request.user.is_superuser:
            return True
            
        if hasattr(obj, 'user'):
            return obj.user == request.user
        return False


class EquipmentPagination(PageNumberPagination):
    page_size = 500
    page_size_query_param = 'page_size'
    max_page_size = 500


class EquipmentViewSet(viewsets.ModelViewSet):
    """
    장비 관리 API (관리자 전용)
    """
    queryset = Equipment.objects.all()
    serializer_class = EquipmentSerializer
    # permission_classes 제거 - 기본 권한 클래스 사용
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        'asset_number', 
        'serial_number', 
        'description', 
        'manufacturer', 
        'model_name', 
        'management_number',
        'rentals__user__username',
        'rentals__user__first_name',
        'rentals__user__last_name'
    ]
    ordering_fields = ['asset_number', 'equipment_type', 'status', 'acquisition_date']
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    pagination_class = EquipmentPagination
    
    def create(self, request, *args, **kwargs):
        """장비 생성 시 상세 로깅 추가"""
        logger = logging.getLogger('rentals')
        logger.info(f"장비 생성 요청 데이터: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            logger.error(f"장비 생성 유효성 검사 오류: {serializer.errors}")
            
            # 오류 메시지를 더 명확하게 처리
            error_messages = {}
            for field, errors in serializer.errors.items():
                if field == 'serial_number' and 'unique' in str(errors):
                    error_messages[field] = ['이미 등록된 일련번호입니다. 다른 일련번호를 입력해주세요.']
                elif field == 'model_name' and 'required' in str(errors):
                    error_messages[field] = ['모델명은 필수 항목입니다.']
                elif field == 'equipment_type' and 'required' in str(errors):
                    error_messages[field] = ['장비 유형은 필수 항목입니다.']
                elif field == 'status' and 'required' in str(errors):
                    error_messages[field] = ['상태는 필수 항목입니다.']
                else:
                    error_messages[field] = errors
            
            return Response(error_messages, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            equipment = serializer.save()
            logger.info(f"장비 생성 성공: {equipment.id}")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            logger.error(f"장비 생성 중 오류: {e}")
            return Response(
                {"detail": f"장비 생성 중 오류가 발생했습니다: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def get_queryset(self):
        """관리자용 장비 목록 조회 시 대여 정보도 함께 로드"""
        queryset = Equipment.objects.prefetch_related(
            Prefetch(
                'rentals',
                queryset=Rental.objects.filter(status='RENTED').select_related('user').order_by('-rental_date'),
                to_attr='current_rentals'
            )
        )
        
        # 상태 필터링
        status = self.request.query_params.get('status', None)
        if status:
            queryset = queryset.filter(status=status)
        
        # 장비 타입 필터링
        equipment_type = self.request.query_params.get('equipment_type', None)
        if equipment_type:
            queryset = queryset.filter(equipment_type=equipment_type)
        
        return queryset

    @action(detail=False, methods=['post'], url_path='update-by-model')
    def update_by_model(self, request):
        """모델명 기준으로 제조사, 생산년도, 구매일시, 구매가격, 장비종류 일괄 업데이트"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        model_name = request.data.get('model_name')
        manufacturer = request.data.get('manufacturer')
        manufacture_year = request.data.get('manufacture_year')
        purchase_date = request.data.get('purchase_date')
        purchase_price = request.data.get('purchase_price')
        equipment_type = request.data.get('equipment_type')
        
        if not model_name:
            return Response({"detail": "모델명이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)
            
        if not manufacturer and not manufacture_year and not purchase_date and purchase_price is None and not equipment_type:
            return Response({"detail": "업데이트할 데이터가 없습니다."}, status=status.HTTP_400_BAD_REQUEST)
        
        # 모델명이 정확히 일치하는 경우
        equipments = Equipment.objects.filter(model_name=model_name)
        
        if not equipments.exists():
            # 모델명이 포함된 장비 검색 (부분 일치)
            equipments = Equipment.objects.filter(model_name__icontains=model_name)
            
        if not equipments.exists():
            return Response({
                "detail": f"모델명 '{model_name}'에 해당하는 장비가 없습니다."
            }, status=status.HTTP_404_NOT_FOUND)
        
        update_data = {}
        
        if manufacturer is not None:
            # 제조사가 빈 문자열인 경우 None으로 설정
            if manufacturer == '':
                update_data['manufacturer'] = None
            else:
                update_data['manufacturer'] = manufacturer
        
        if manufacture_year is not None:
            try:
                manufacture_year = int(manufacture_year)
                update_data['manufacture_year'] = manufacture_year
            except (ValueError, TypeError):
                return Response({
                    "detail": "생산년도는 유효한 연도(숫자)여야 합니다."
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if purchase_date is not None:
            try:
                # purchase_date가 문자열로 왔다면 date로 변환 시도
                if isinstance(purchase_date, str):
                    from datetime import datetime
                    parsed_date = datetime.strptime(purchase_date, '%Y-%m-%d').date()
                    purchase_date = parsed_date
                elif isinstance(purchase_date, datetime):
                    # datetime 객체인 경우 date로 변환
                    purchase_date = purchase_date.date()
                update_data['purchase_date'] = purchase_date
            except Exception as e:
                return Response({
                    "detail": f"구매일 처리 중 오류가 발생했습니다. YYYY-MM-DD 형식으로 입력해주세요: {str(e)}"
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if purchase_price is not None:
            try:
                # 빈 문자열이나 None이 아닌 경우에만 처리
                if purchase_price != '' and purchase_price is not None:
                    purchase_price = float(purchase_price)
                    if purchase_price < 0:
                        return Response({
                            "detail": "구매가격은 0 이상이어야 합니다."
                        }, status=status.HTTP_400_BAD_REQUEST)
                    update_data['purchase_price'] = purchase_price
                elif purchase_price == '':
                    # 빈 문자열인 경우 None으로 설정 (필드 초기화)
                    update_data['purchase_price'] = None
            except (ValueError, TypeError):
                return Response({
                    "detail": "구매가격은 유효한 숫자여야 합니다."
                }, status=status.HTTP_400_BAD_REQUEST)
        
        if equipment_type is not None and equipment_type != '':
            # 장비 종류 유효성 검사
            valid_types = [choice[0] for choice in Equipment.EQUIPMENT_TYPE_CHOICES]
            if equipment_type not in valid_types:
                return Response({
                    "detail": f"유효하지 않은 장비 종류입니다. 사용 가능한 종류: {', '.join(valid_types)}"
                }, status=status.HTTP_400_BAD_REQUEST)
            update_data['equipment_type'] = equipment_type
        
        # 일괄 업데이트 수행
        count = equipments.count()
        updated_equipments = []
        failed_equipments = []
        
        # 로깅 추가
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"일괄 업데이트 시작: 모델 '{model_name}', 총 {count}개 장비")
        
        # 일괄처리 시 모든 장비의 관리번호를 완전히 새로 부여하기 위해 순서대로 처리
        equipments_list = list(equipments.order_by('created_at'))
        
        # 각 장비를 개별적으로 업데이트하여 관리번호 중복 방지
        for index, equipment in enumerate(equipments_list):
            try:
                logger.info(f"장비 {equipment.id} (모델: {equipment.model_name}) 업데이트 시작 - 순서: {index + 1}/{count}")
                
                # 기존 값 백업
                old_equipment_type = equipment.equipment_type
                old_purchase_date = equipment.purchase_date
                old_management_number = equipment.management_number
                
                # 일괄 업데이트 시에는 모든 장비의 관리번호를 강제로 재생성
                logger.info(f"장비 {equipment.id}: 일괄 업데이트로 인한 관리번호 재생성")
                equipment.management_number = None
                
                # equipment_type이 변경되는 경우 추가 처리
                if 'equipment_type' in update_data and equipment.equipment_type != update_data['equipment_type']:
                    logger.info(f"장비 {equipment.id}: equipment_type 변경 ({equipment.equipment_type} -> {update_data['equipment_type']})")
                    # 데이터베이스에서 다시 조회하여 최신 상태 가져오기
                    equipment.refresh_from_db()
                
                # 업데이트할 필드들을 설정
                if 'manufacturer' in update_data:
                    equipment.manufacturer = update_data['manufacturer']
                    logger.info(f"장비 {equipment.id}: manufacturer 업데이트 -> {update_data['manufacturer']}")
                if 'manufacture_year' in update_data:
                    equipment.manufacture_year = update_data['manufacture_year']
                    logger.info(f"장비 {equipment.id}: manufacture_year 업데이트 -> {update_data['manufacture_year']}")
                if 'purchase_date' in update_data:
                    # 구매일시가 변경되는 경우
                    equipment.purchase_date = update_data['purchase_date']
                    logger.info(f"장비 {equipment.id}: purchase_date 업데이트 -> {update_data['purchase_date']}")
                if 'purchase_price' in update_data:
                    equipment.purchase_price = update_data['purchase_price']
                    logger.info(f"장비 {equipment.id}: purchase_price 업데이트 -> {update_data['purchase_price']}")
                if 'equipment_type' in update_data:
                    equipment.equipment_type = update_data['equipment_type']
                    logger.info(f"장비 {equipment.id}: equipment_type 업데이트 -> {update_data['equipment_type']}")
                
                # 일괄처리 시 순서대로 관리번호 부여 (1번부터 연속)
                if equipment.purchase_date:
                    # 장비 유형 이니셜 가져오기
                    initial = equipment.EQUIPMENT_TYPE_INITIALS.get(equipment.equipment_type, 'O')
                    year = equipment.purchase_date.year
                    
                    if equipment.model_name:
                        # EquipmentModel에서 모델 번호 가져오기
                        try:
                            from rentals.models import EquipmentModel
                            model_number = EquipmentModel.get_or_create_model_number(
                                equipment.equipment_type, 
                                year, 
                                equipment.model_name
                            )
                        except Exception as e:
                            logger.error(f"EquipmentModel 생성 중 오류: {str(e)}")
                            model_number = 1
                        
                        # 순서대로 관리번호 부여 (1번부터 연속)
                        sequence = index + 1
                        base_number = f"{initial}-{year}-{model_number:02d}"
                        equipment.management_number = f"{base_number}-{sequence:03d}"
                        
                        logger.info(f"장비 {equipment.id}: 순서 {sequence}번으로 관리번호 부여 -> {equipment.management_number}")
                    else:
                        # 모델명이 없는 경우
                        base_number = f"{initial}-{year}"
                        sequence = index + 1
                        equipment.management_number = f"{base_number}-{sequence:03d}"
                        
                        logger.info(f"장비 {equipment.id}: 순서 {sequence}번으로 관리번호 부여 -> {equipment.management_number}")
                
                # save() 메서드를 호출하여 관리번호 업데이트 로직이 작동하도록 함
                try:
                    equipment.save()
                    logger.info(f"장비 {equipment.id} 업데이트 완료, 새로운 관리번호: {equipment.management_number}")
                    updated_equipments.append(equipment)
                except Exception as save_error:
                    # save() 중 오류가 발생한 경우 관리번호를 수동으로 처리
                    logger.warning(f"장비 {equipment.id} save() 중 오류: {str(save_error)}")
                    
                    # 관리번호가 중복인 경우 수동으로 해결
                    if 'management_number' in str(save_error).lower() or 'unique' in str(save_error).lower():
                        logger.info(f"장비 {equipment.id} 관리번호 중복 문제 해결 시도")
                        
                        # 기존 관리번호를 None으로 설정
                        equipment.management_number = None
                        equipment.save(update_fields=['management_number'])
                        
                        # 새로운 관리번호 생성
                        try:
                            new_management_number = equipment.generate_management_number()
                            if new_management_number:
                                equipment.management_number = new_management_number
                                equipment.save(update_fields=['management_number'])
                                logger.info(f"장비 {equipment.id} 관리번호 재생성 성공: {new_management_number}")
                                updated_equipments.append(equipment)
                            else:
                                logger.error(f"장비 {equipment.id} 관리번호 생성 실패")
                                raise save_error
                        except Exception as retry_error:
                            logger.error(f"장비 {equipment.id} 관리번호 재생성 실패: {str(retry_error)}")
                            raise retry_error
                    else:
                        # 다른 오류인 경우 원래 오류를 다시 발생시킴
                        raise save_error
                
            except Exception as e:
                # 개별 장비 처리 중 오류 발생 시 로그 기록 및 실패 목록에 추가
                import traceback
                error_traceback = traceback.format_exc()
                
                logger.error(f"장비 {equipment.id} (모델: {equipment.model_name}) 업데이트 중 오류: {str(e)}")
                logger.error(f"장비 상세 정보: equipment_type={equipment.equipment_type}, purchase_date={equipment.purchase_date}, management_number={equipment.management_number}")
                logger.error(f"업데이트 데이터: {update_data}")
                logger.error(f"에러 상세 정보: {error_traceback}")
                
                failed_equipments.append({
                    'id': equipment.id,
                    'model_name': equipment.model_name,
                    'equipment_type': equipment.equipment_type,
                    'purchase_date': str(equipment.purchase_date) if equipment.purchase_date else None,
                    'management_number': equipment.management_number,
                    'error': str(e),
                    'error_traceback': error_traceback
                })
                continue
        
        # 업데이트된 장비 목록 반환
        serializer = self.get_serializer(updated_equipments, many=True)
        
        # 성공/실패 메시지 구성
        success_count = len(updated_equipments)
        failed_count = len(failed_equipments)
        
        logger.info(f"일괄 업데이트 완료: 성공 {success_count}개, 실패 {failed_count}개")
        
        if failed_count == 0:
            message = f"{success_count}개의 장비가 성공적으로 업데이트되었습니다."
        else:
            message = f"{success_count}개의 장비가 업데이트되었고, {failed_count}개의 장비에서 오류가 발생했습니다."
        
        response_data = {
            "success": True,
            "message": message,
            "updated_count": success_count,
            "failed_count": failed_count,
            "updated_equipments": serializer.data
        }
        
        if failed_equipments:
            response_data["failed_equipments"] = failed_equipments
        
        return Response(response_data)

    @action(detail=False, methods=['get'], url_path='get-model-info')
    def get_model_info(self, request):
        """모델명에 해당하는 장비의 정보를 가져옴"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        model_name = request.query_params.get('model_name')
        if not model_name:
            return Response({"detail": "모델명이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)
            
        # 모델명이 정확히 일치하는 장비들 중에서 가장 많이 사용된 장비 유형을 찾음
        from django.db.models import Count
        
        # 정확히 일치하는 모델명의 장비들
        equipments = Equipment.objects.filter(model_name=model_name)
        
        if not equipments.exists():
            # 모델명이 포함된 장비 검색 (부분 일치)
            equipments = Equipment.objects.filter(model_name__icontains=model_name)
            
        if not equipments.exists():
            return Response({
                "detail": f"모델명 '{model_name}'에 해당하는 장비가 없습니다."
            }, status=status.HTTP_404_NOT_FOUND)
        
        # 가장 많이 사용된 장비 유형 찾기
        equipment_type_counts = equipments.values('equipment_type').annotate(count=Count('id')).order_by('-count')
        most_common_type = equipment_type_counts.first()['equipment_type'] if equipment_type_counts else None
        
        # 가장 먼저 등록된 장비의 다른 정보들 가져오기
        first_equipment = equipments.order_by('created_at').first()
        
        return Response({
            "success": True,
            "data": {
                "model_name": first_equipment.model_name,
                "manufacturer": first_equipment.manufacturer,
                "manufacture_year": first_equipment.manufacture_year,
                "purchase_date": first_equipment.purchase_date,
                "purchase_price": first_equipment.purchase_price,
                "equipment_type": most_common_type or first_equipment.equipment_type
            }
        })

    def update(self, request, *args, **kwargs):
        """장비 수정 API"""
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f"EquipmentViewSet.update 호출됨 - ID: {kwargs.get('pk')}")
        logger.info(f"요청 데이터: {request.data}")
        logger.info(f"요청 사용자: {request.user}, 관리자 여부: {request.user.is_superuser}")
        
        # 제조사 정보 로깅 추가
        if 'manufacturer' in request.data:
            logger.info(f"제조사 업데이트 요청: {request.data.get('manufacturer')}")
        
        # 인증 확인
        if not request.user.is_authenticated:
            logger.warning(f"인증되지 않은 사용자가 장비 수정 시도")
            return Response(
                {"detail": "인증이 필요합니다."}, 
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        # 관리자 권한 확인
        if not request.user.is_superuser:
            logger.warning(f"권한 없는 사용자가 장비 수정 시도 - 사용자: {request.user.username}")
            return Response(
                {"detail": "관리자 권한이 필요합니다."}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            instance = self.get_object()
            logger.info(f"기존 장비 정보: ID={instance.id}, 시리얼={instance.serial_number}")
            
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            
            if not serializer.is_valid():
                logger.error(f"유효성 검사 실패: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            # 장비 상태 변경 처리
            old_status = instance.status
            self.perform_update(serializer)
            new_status = serializer.instance.status
            
            # 상태 변경 공통 처리(서비스 호출)
            from .services.rental_logic import change_equipment_status
            auto_count = 0
            if new_status in ['AVAILABLE', 'RENTED']:
                renter = None
                due_date = None
                notes = ''
                if new_status == 'RENTED':
                    rental_data = request.data.get('rental')
                    if rental_data and rental_data.get('user_id'):
                        from django.contrib.auth import get_user_model
                        User = get_user_model()
                        renter = User.objects.get(id=rental_data['user_id'])
                        due_date = rental_data.get('due_date')
                        notes = rental_data.get('notes', '')
                auto_count, created_rental = change_equipment_status(
                    equipment=instance,
                    new_status=new_status,
                    actor=request.user,
                    reason=request.data.get('reason', ''),
                    renter=renter,
                    due_date=due_date,
                    notes=notes,
                )
                logger.info(f"장비 {instance.id} 상태 변경 처리 완료: {old_status} -> {new_status} (자동 반납 {auto_count}건)")
            
            # 기존 개별 대여 처리 블록 제거(서비스로 대체)
            
            logger.info(f"장비 업데이트 성공 - ID: {instance.id}")
            
            # 업데이트된 장비 정보 로깅
            try:
                updated_equipment = Equipment.objects.get(pk=instance.id)
                logger.info(f"업데이트된 장비 정보: 제조사={updated_equipment.manufacturer}, 모델={updated_equipment.model_name}, 시리얼={updated_equipment.serial_number}")
            except Equipment.DoesNotExist:
                logger.warning(f"업데이트된 장비 정보를 가져올 수 없음 - ID: {instance.id}")
            
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"장비 업데이트 중 예외 발생: {str(e)}")
            return Response(
                {'error': f'장비 업데이트 중 오류가 발생했습니다: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['post'], parser_classes=[JSONParser, MultiPartParser, FormParser])
    def register_maintenance(self, request):
        """자동 설치를 위한 장비 등록 (유지보수 상태)"""
        logger = logging.getLogger(__name__)
        logger.warning("REGISTER_MAINTENANCE 메서드가 호출되었습니다!")
        logger.warning(f"요청 데이터: {request.data}")
        logger.warning(f"요청 헤더: {dict(request.headers)}")
        logger.warning(f"요청 메서드: {request.method}")
        logger.warning(f"요청 경로: {request.path}")
        logger.warning(f"Content-Type: {request.content_type}")
        
        # 시리얼 번호 체크
        serial_number = request.data.get('serial_number')
        if serial_number:
            try:
                existing_equipment = Equipment.objects.get(serial_number=serial_number)
                logger.warning(f"이미 등록된 시리얼 번호: {serial_number}")
                return Response({
                    'error': '이미 등록된 시리얼 번호입니다.',
                    'existing_equipment': {
                        'id': existing_equipment.id,
                        'asset_number': existing_equipment.asset_number,
                        'equipment_type': existing_equipment.get_equipment_type_display(),
                        'status': existing_equipment.get_status_display()
                    }
                }, status=status.HTTP_409_CONFLICT)
            except Equipment.DoesNotExist:
                logger.info(f"새로운 시리얼 번호: {serial_number}")

        # MAC 주소 목록이 전달되었는지 확인
        mac_addresses = request.data.get('mac_addresses', [])
        if not mac_addresses:
            logger.error("MAC 주소 배열이 없습니다.")
            return Response(
                {'error': 'mac_addresses 배열이 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 장비 정보 준비
        equipment_data = request.data.copy()
        
        # 자동 설치용 장비는 유지보수 상태로 설정
        equipment_data['status'] = 'MAINTENANCE'
        # acquisition_date를 현재 날짜로 설정 (DateField이므로 date 객체 필요)
        equipment_data['acquisition_date'] = timezone.now().date()
        
        logger.warning(f"장비 등록 시도 데이터: {equipment_data}")
        
        # 장비 등록
        serializer = self.get_serializer(data=equipment_data)
        if serializer.is_valid():
            equipment = serializer.save()
            
            logger.info(f"자동 설치용 장비 등록 성공: {equipment.id} (상태: MAINTENANCE)")
            return Response({
                "success": True,
                "data": serializer.data,
                "message": "자동 설치용 장비가 유지보수 상태로 등록되었습니다."
            }, status=status.HTTP_201_CREATED)
        
        logger.error(f"장비 등록 실패 - 유효성 검사 오류: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='report-status', parser_classes=[JSONParser, MultiPartParser, FormParser])
    def report_status(self, request):
        """장비 상태 수집 전용(비로그인 허용) - 기존 유지보수 등록 로직 재사용"""
        logger = logging.getLogger(__name__)
        logger.info("report_status 호출 - register_maintenance 로직 위임")
        # 기존 자동 유지보수 등록 로직을 그대로 사용하여 등록 및 상태 반영
        return self.register_maintenance(request)

    @action(detail=False, methods=['post'], url_path='import')
    def import_excel(self, request):
        """엑셀 파일로부터 장비를 일괄 추가하는 API (관리자 전용)"""
        if not request.user.is_superuser:
            return Response({"detail": "관리자 권한이 필요합니다."}, status=status.HTTP_403_FORBIDDEN)
        
        if 'file' not in request.FILES:
            return Response({'error': '파일이 제공되지 않았습니다.'}, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        
        try:
            # 엑셀 파일 읽기
            df = pd.read_excel(excel_file)
            
            # 필수 열 확인
            required_columns = ['물품번호', '장비 유형', '시리얼 번호', '취득일']
            for col in required_columns:
                if col not in df.columns:
                    return Response({'error': f'필수 열 "{col}"이 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 장비 생성
            created_equipment = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # 취득일 처리
                    acquisition_date = row.get('취득일')
                    if pd.isnull(acquisition_date):
                        acquisition_date = timezone.now().date()
                    elif isinstance(acquisition_date, str):
                        try:
                            acquisition_date = pd.to_datetime(acquisition_date).date()
                        except:
                            acquisition_date = timezone.now().date()
                    
                    equipment_data = {
                        'asset_number': row['물품번호'],
                        'equipment_type': row['장비 유형'],
                        'serial_number': row['시리얼 번호'],
                        'description': row.get('설명', ''),
                        'status': row.get('장비 상태', 'AVAILABLE'),
                        'acquisition_date': acquisition_date
                    }
                    
                    # 중복 확인
                    if Equipment.objects.filter(serial_number=equipment_data['serial_number']).exists():
                        errors.append({
                            'row': index + 2,  # 엑셀 행 번호 (헤더 + 1)
                            'asset_number': equipment_data['asset_number'],
                            'errors': '이미 존재하는 시리얼 번호입니다.'
                        })
                        continue
                    
                    # 장비 생성
                    serializer = EquipmentSerializer(data=equipment_data)
                    if serializer.is_valid():
                        serializer.save()
                        created_equipment.append(equipment_data['asset_number'])
                    else:
                        errors.append({
                            'row': index + 2,
                            'asset_number': equipment_data['asset_number'],
                            'errors': serializer.errors
                        })
                except Exception as e:
                    errors.append({
                        'row': index + 2,
                        'asset_number': row.get('물품번호', '알 수 없음'),
                        'errors': str(e)
                    })
            
            return Response({
                'success': True,
                'created_equipment': created_equipment,
                'errors': errors,
                'total_created': len(created_equipment),
                'total_errors': len(errors)
            })
            
        except Exception as e:
            return Response({'error': f'파일 처리 중 오류가 발생했습니다: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], parser_classes=[JSONParser, MultiPartParser, FormParser])
    def register(self, request):
        """MAC 주소로 장비를 등록하고 대여 정보도 생성"""
        logger = logging.getLogger(__name__)
        
        # 시리얼 번호 체크
        serial_number = request.data.get('serial_number')
        if serial_number:
            try:
                existing_equipment = Equipment.objects.get(serial_number=serial_number)
                current_rental = existing_equipment.rentals.filter(status='RENTED').first()
                rental_info = None
                if current_rental and current_rental.user:
                    rental_info = {
                        'user': {
                            'username': current_rental.user.username,
                            'name': current_rental.user.get_full_name() or current_rental.user.username
                        }
                    }
                
                return Response({
                    'error': '이미 등록된 시리얼 번호입니다.',
                    'existing_equipment': {
                        'id': existing_equipment.id,
                        'asset_number': existing_equipment.asset_number,
                        'equipment_type': existing_equipment.get_equipment_type_display(),
                        'status': existing_equipment.get_status_display(),
                        'rental': rental_info
                    }
                }, status=status.HTTP_400_BAD_REQUEST)
            except Equipment.DoesNotExist:
                pass

        # MAC 주소 목록이 전달되었는지 확인
        mac_addresses = request.data.get('mac_addresses', [])
        if not mac_addresses:
            return Response(
                {'error': 'mac_addresses 배열이 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # MAC 주소로 사용자 찾기
        user = None
        for mac_item in mac_addresses:
            mac_address = mac_item.get('mac_address') if isinstance(mac_item, dict) else mac_item
            if not mac_address:
                continue
                
            try:
                # Device 모델에서 MAC 주소로 사용자 찾기
                device = Device.objects.get(mac_address=mac_address)
                if device.user:
                    user = device.user
                    break
            except Device.DoesNotExist:
                continue
        
        if not user:
            return Response(
                {'error': '등록된 MAC 주소를 가진 사용자를 찾을 수 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # 장비 정보 준비
        equipment_data = request.data.copy()
        
        # 장비 등록
        serializer = self.get_serializer(data=equipment_data)
        if serializer.is_valid():
            equipment = serializer.save()
            
            # 대여 정보 생성
            rental_info = None
            try:
                # 장비 상태를 '대여 중'으로 변경
                equipment.status = 'RENTED'
                equipment.save()
                
                # 대여 정보 생성
                rental = Rental.objects.create(
                    user=user,
                    equipment=equipment,
                    rental_date=timezone.now(),
                    due_date=timezone.now() + timezone.timedelta(days=30),  # 30일 후 반납 예정
                    status='RENTED',
                    notes=f'장비 등록 시 자동 생성된 대여 정보',
                    approved_by=user  # 사용자가 자신의 장비를 등록하므로 자동 승인
                )
                
                # 결과에 포함할 대여 정보
                rental_info = RentalSerializer(rental).data
                print(f"대여 정보 생성 성공: {rental.id}")
            except Exception as e:
                print(f"대여 정보 생성 중 오류 발생: {str(e)}")
            
            # 결과 반환
            result = serializer.data
            if rental_info:
                result['rental_info'] = rental_info
            
            logger.info(f"대여 생성 성공: {rental.id}")
            return Response({
                "success": True,
                "data": result,
                "message": "대여가 성공적으로 생성되었습니다."
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='change-status')
    def change_status(self, request, pk=None):
        """장비 상태 변경 API"""
        logger = logging.getLogger('rentals')
        if not request.user.is_superuser:
            return Response({"detail": "관리자 권한이 필요합니다."}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            equipment = self.get_object()
            new_status = request.data.get('status')
            reason = request.data.get('reason', '')
            
            if not new_status:
                return Response({"detail": "새로운 상태를 지정해주세요."}, status=status.HTTP_400_BAD_REQUEST)
            
            # 유효한 상태인지 확인
            valid_statuses = ['AVAILABLE', 'RENTED', 'MAINTENANCE', 'BROKEN', 'LOST', 'DISPOSED']
            if new_status not in valid_statuses:
                return Response({"detail": f"유효하지 않은 상태입니다. 가능한 상태: {', '.join(valid_statuses)}"}, status=status.HTTP_400_BAD_REQUEST)
            
            old_status = equipment.status
            if old_status == new_status:
                return Response({"detail": "현재 상태와 동일합니다."}, status=status.HTTP_400_BAD_REQUEST)
            
            # 상태 변경 및 부가 처리
            with transaction.atomic():
                # 공통 서비스 사용
                from .services.rental_logic import change_equipment_status
                renter = None
                due_date = None
                notes = ''
                if new_status == 'RENTED' and request.data.get('user_id'):
                    from users.models import User
                    renter = User.objects.get(id=request.data['user_id'])
                    due_date = request.data.get('due_date')
                    notes = request.data.get('notes', '')

                auto_count, created_rental = change_equipment_status(
                    equipment=equipment,
                    new_status=new_status,
                    actor=request.user,
                    reason=reason,
                    renter=renter,
                    due_date=due_date,
                    notes=notes,
                )

                # 상태 저장은 서비스 내에서 처리됨
            
            # 상태 표시명 매핑
            status_display_map = {
                'AVAILABLE': '사용 가능',
                'RENTED': '대여 중',
                'MAINTENANCE': '유지보수',
                'BROKEN': '파손',
                'LOST': '분실',
                'DISPOSED': '폐기'
            }
            
            old_status_display = status_display_map.get(old_status, old_status)
            new_status_display = status_display_map.get(new_status, new_status)

            # 자동 반납 건수 표시를 위해 details 문자열 생성 시점을 변경
            # auto_returned_count는 AVAILABLE 전환 시 활성 대여 자동 반납 개수
            details = f"장비 상태 변경: {old_status_display} → {new_status_display}"
            # 자동 반납이 있었다면 괄호로 개수 표기 (아래 블록에서 count 증가)
            # reason은 항상 마지막에 추가
            
            # reason 및 자동 반납 건수를 반영하기 위해 다시 details 구성
            try:
                # auto_returned_count가 존재하면 로컬 변수에서 참조되도록
                auto_count = locals().get('auto_returned_count', 0)
                if auto_count and auto_count > 0:
                    details += f" (자동 반납 {auto_count}건)"
            except Exception:
                pass
            if reason:
                details += f" (사유: {reason})"

            EquipmentHistory.objects.create(
                equipment=equipment,
                action='STATUS_CHANGED',
                user=request.user,
                old_value={'status': old_status},
                new_value={'status': new_status, 'reason': reason},
                details=details
            )
            
            logger.info(f"장비 상태 변경: {equipment.asset_number} {old_status} → {new_status} by {request.user.username}")
            
            return Response({
                "success": True,
                "message": f"장비 상태가 {old_status_display}에서 {new_status_display}로 변경되었습니다.",
                "old_status": old_status,
                "new_status": new_status,
                "details": details
            })
            
        except Exception as e:
            logger.error(f"장비 상태 변경 중 오류: {e}")
            return Response({"detail": f"장비 상태 변경 중 오류가 발생했습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='history')
    def get_equipment_history(self, request, pk=None):
        """특정 장비의 이력 조회 - 모든 인증된 사용자 접근 가능"""
        logger = logging.getLogger(__name__)
        
        logger.info(f"=== 장비 이력 조회 API 호출됨 ===")
        logger.info(f"요청 URL: {request.path}")
        logger.info(f"요청 메서드: {request.method}")
        logger.info(f"사용자: {request.user.username}")
        logger.info(f"장비 ID: {pk}")
        
        try:
            # 장비 존재 여부 확인
            try:
                equipment = self.get_object()
                logger.info(f"장비 조회 성공: ID={equipment.id}, 관리번호={equipment.asset_number}")
            except Exception as e:
                logger.error(f"장비 조회 실패: {e}")
                return Response({"detail": f"장비를 찾을 수 없습니다: {str(e)}"}, status=status.HTTP_404_NOT_FOUND)
            
            # EquipmentHistory 모델 import 확인
            try:
                from .models import EquipmentHistory
                logger.info("EquipmentHistory 모델 import 성공")
            except Exception as e:
                logger.error(f"EquipmentHistory 모델 import 실패: {e}")
                return Response({"detail": f"EquipmentHistory 모델을 찾을 수 없습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # EquipmentHistorySerializer import 확인
            try:
                from .serializers import EquipmentHistorySerializer
                logger.info("EquipmentHistorySerializer import 성공")
            except Exception as e:
                logger.error(f"EquipmentHistorySerializer import 실패: {e}")
                return Response({"detail": f"EquipmentHistorySerializer를 찾을 수 없습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # 장비 이력 조회
            try:
                history = EquipmentHistory.objects.filter(equipment=equipment).order_by('-created_at')
                history_count = history.count()
                logger.info(f"조회된 이력 개수: {history_count}")
            except Exception as e:
                logger.error(f"이력 조회 실패: {e}")
                return Response({"detail": f"이력을 조회할 수 없습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # 이력이 없는 경우 로그 출력
            if history_count == 0:
                logger.warning(f"장비 ID={equipment.id}에 대한 이력이 없습니다.")
                # 장비 생성 이력이 없는 경우 자동으로 생성
                try:
                    if not EquipmentHistory.objects.filter(equipment=equipment, action='CREATED').exists():
                        logger.info(f"장비 ID={equipment.id}에 대한 생성 이력을 자동으로 생성합니다.")
                        EquipmentHistory.objects.create(
                            equipment=equipment,
                            action='CREATED',
                            user=request.user,
                            new_value={
                                'asset_number': equipment.asset_number,
                                'manufacturer': equipment.manufacturer,
                                'model_name': equipment.model_name,
                                'equipment_type': equipment.equipment_type,
                                'serial_number': equipment.serial_number,
                                'status': equipment.status,
                                'acquisition_date': equipment.acquisition_date.isoformat() if equipment.acquisition_date else None,
                                'manufacture_year': equipment.manufacture_year,
                                'purchase_date': equipment.purchase_date.isoformat() if equipment.purchase_date else None,
                                'purchase_price': str(equipment.purchase_price) if equipment.purchase_price else None,
                                'management_number': equipment.management_number
                            },
                            details=f"자동 생성: 장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 생성"
                        )
                        # 다시 이력 조회
                        history = EquipmentHistory.objects.filter(equipment=equipment).order_by('-created_at')
                        history_count = history.count()
                        logger.info(f"자동 생성 후 이력 개수: {history_count}")
                except Exception as e:
                    logger.error(f"자동 이력 생성 실패: {e}")
                    return Response({"detail": f"이력을 생성할 수 없습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # 시리얼라이저 처리
            try:
                serializer = EquipmentHistorySerializer(history, many=True)
                logger.info(f"시리얼라이저 처리 성공")
            except Exception as e:
                logger.error(f"시리얼라이저 처리 실패: {e}")
                return Response({"detail": f"이력 데이터를 처리할 수 없습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            logger.info(f"장비 이력 조회 완료: {history_count}건 반환")
            logger.info(f"=== 장비 이력 조회 API 완료 ===")
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"장비 이력 조회 중 예상치 못한 오류: {e}")
            import traceback
            logger.error(f"상세 오류: {traceback.format_exc()}")
            logger.error(f"=== 장비 이력 조회 API 오류 ===")
            return Response({"detail": f"장비 이력 조회 중 오류가 발생했습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def perform_create(self, serializer):
        """장비 생성 시 이력 기록"""
        logger = logging.getLogger(__name__)
        
        equipment = serializer.save()
        
        # 이력 기록
        EquipmentHistory.objects.create(
            equipment=equipment,
            action='CREATED',
            user=self.request.user,
            new_value={
                'asset_number': equipment.asset_number,
                'manufacturer': equipment.manufacturer,
                'model_name': equipment.model_name,
                'equipment_type': equipment.equipment_type,
                'serial_number': equipment.serial_number,
                'status': equipment.status,
                'acquisition_date': equipment.acquisition_date.isoformat() if equipment.acquisition_date else None,
                'manufacture_year': equipment.manufacture_year,
                'purchase_date': equipment.purchase_date.isoformat() if equipment.purchase_date else None,
                'purchase_price': str(equipment.purchase_price) if equipment.purchase_price else None,
                'management_number': equipment.management_number
            },
            details=f"장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 생성"
        )
        
        logger.info(f"장비 생성: {equipment.asset_number} by {self.request.user.username}")

    def perform_update(self, serializer):
        """장비 수정 시 이력 기록"""
        logger = logging.getLogger(__name__)
        
        # 관리자 권한 확인
        if not self.request.user.is_superuser:
            raise ValidationError({"detail": "관리자 권한이 필요합니다."})
        
        old_instance = Equipment.objects.get(pk=serializer.instance.pk)
        old_data = {
            'asset_number': old_instance.asset_number,
            'manufacturer': old_instance.manufacturer,
            'model_name': old_instance.model_name,
            'equipment_type': old_instance.equipment_type,
            'serial_number': old_instance.serial_number,
            'status': old_instance.status,
            'acquisition_date': old_instance.acquisition_date.isoformat() if old_instance.acquisition_date else None,
            'manufacture_year': old_instance.manufacture_year,
            'purchase_date': old_instance.purchase_date.isoformat() if old_instance.purchase_date else None,
            'purchase_price': str(old_instance.purchase_price) if old_instance.purchase_price else None,
            'management_number': old_instance.management_number
        }
        
        equipment = serializer.save()
        
        new_data = {
            'asset_number': equipment.asset_number,
            'manufacturer': equipment.manufacturer,
            'model_name': equipment.model_name,
            'equipment_type': equipment.equipment_type,
            'serial_number': equipment.serial_number,
            'status': equipment.status,
            'acquisition_date': equipment.acquisition_date.isoformat() if equipment.acquisition_date else None,
            'manufacture_year': equipment.manufacture_year,
            'purchase_date': equipment.purchase_date.isoformat() if equipment.purchase_date else None,
            'purchase_price': str(equipment.purchase_price) if equipment.purchase_price else None,
            'management_number': equipment.management_number
        }
        
        # 상태 변경 여부 확인
        status_changed = old_data['status'] != new_data['status']
        action = 'STATUS_CHANGED' if status_changed else 'UPDATED'
        
        # 상태 변경인 경우 더 자세한 이력 기록
        if status_changed:
            status_display_map = {
                'AVAILABLE': '사용 가능',
                'RENTED': '대여 중',
                'MAINTENANCE': '유지보수',
                'BROKEN': '파손',
                'LOST': '분실',
                'DISPOSED': '폐기'
            }
            old_status_display = status_display_map.get(old_data['status'], old_data['status'])
            new_status_display = status_display_map.get(new_data['status'], new_data['status'])
            
            details = f"장비 상태 변경: {old_status_display} → {new_status_display}"
        else:
            details = f"장비 정보 수정: '{equipment.asset_number or equipment.model_name or equipment.serial_number}'"
        
        # 이력 기록
        EquipmentHistory.objects.create(
            equipment=equipment,
            action=action,
            user=self.request.user,
            old_value=old_data,
            new_value=new_data,
            details=details
        )
        
        logger.info(f"장비 수정: {equipment.asset_number} by {self.request.user.username} - {details}")

    def perform_destroy(self, instance):
        """장비 삭제 시 이력 기록"""
        logger = logging.getLogger(__name__)
        
        # 삭제 이력 기록
        EquipmentHistory.objects.create(
            equipment=instance,
            action='DELETED',
            user=self.request.user,
            old_value={
                'asset_number': instance.asset_number,
                'manufacturer': instance.manufacturer,
                'model_name': instance.model_name,
                'equipment_type': instance.equipment_type,
                'serial_number': instance.serial_number,
                'status': instance.status,
                'acquisition_date': instance.acquisition_date.isoformat() if instance.acquisition_date else None,
                'manufacture_year': instance.manufacture_year,
                'purchase_date': instance.purchase_date.isoformat() if instance.purchase_date else None,
                'purchase_price': str(instance.purchase_price) if instance.purchase_price else None,
                'management_number': instance.management_number
            },
            details=f"장비 '{instance.asset_number or instance.model_name or instance.serial_number}' 삭제"
        )
        
        logger.info(f"장비 삭제: {instance.asset_number} by {self.request.user.username}")
        instance.delete()

    def destroy(self, request, *args, **kwargs):
        """장비 삭제 후 성공 메시지 반환"""
        instance = self.get_object()
        asset_number = instance.asset_number or instance.model_name or instance.serial_number
        
        try:
            self.perform_destroy(instance)
            return Response({
                'message': f"장비 '{asset_number}'이(가) 성공적으로 삭제되었습니다."
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'error': f"장비 '{asset_number}' 삭제 중 오류가 발생했습니다.",
                'details': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='batch-change-status')
    def batch_change_status(self, request):
        """여러 장비의 상태를 한번에 변경하는 API"""
        logger = logging.getLogger('rentals')
        
        if not request.user.is_superuser:
            return Response({"detail": "관리자 권한이 필요합니다."}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            equipment_ids = request.data.get('equipment_ids', [])
            new_status = request.data.get('status')
            reason = request.data.get('reason', '')
            user_id = request.data.get('user_id')  # 대여중 상태일 때 사용자 ID
            
            if not equipment_ids:
                return Response({"detail": "장비 ID 목록을 제공해주세요."}, status=status.HTTP_400_BAD_REQUEST)
            
            if not new_status:
                return Response({"detail": "새로운 상태를 지정해주세요."}, status=status.HTTP_400_BAD_REQUEST)
            
            # 유효한 상태인지 확인
            valid_statuses = ['AVAILABLE', 'RENTED', 'MAINTENANCE', 'BROKEN', 'LOST', 'DISPOSED']
            if new_status not in valid_statuses:
                return Response({"detail": f"유효하지 않은 상태입니다. 가능한 상태: {', '.join(valid_statuses)}"}, status=status.HTTP_400_BAD_REQUEST)
            
            # 대여중 상태일 때 사용자 확인
            if new_status == 'RENTED' and not user_id:
                return Response({"detail": "대여중 상태로 변경할 때는 사용자를 지정해주세요."}, status=status.HTTP_400_BAD_REQUEST)
            
            # 사용자 조회 (대여중 상태일 때)
            user = None
            if new_status == 'RENTED' and user_id:
                try:
                    from users.models import User
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return Response({"detail": "지정된 사용자를 찾을 수 없습니다."}, status=status.HTTP_404_NOT_FOUND)
            
            # 상태 표시명 매핑
            status_display_map = {
                'AVAILABLE': '사용 가능',
                'RENTED': '대여 중',
                'MAINTENANCE': '유지보수',
                'BROKEN': '파손',
                'LOST': '분실',
                'DISPOSED': '폐기'
            }
            new_status_display = status_display_map.get(new_status, new_status)
            
            # 장비 조회
            equipments = Equipment.objects.filter(id__in=equipment_ids)
            if not equipments.exists():
                return Response({"detail": "지정된 장비를 찾을 수 없습니다."}, status=status.HTTP_404_NOT_FOUND)
            
            success_count = 0
            failed_count = 0
            errors = []
            
            with transaction.atomic():
                for equipment in equipments:
                    try:
                        old_status = equipment.status
                        if old_status == new_status:
                            errors.append(f"장비 {equipment.asset_number or equipment.model_name}: 이미 {new_status_display} 상태입니다.")
                            failed_count += 1
                            continue
                        
                        # AVAILABLE로 변경 시 활성 대여 자동 반납 처리
                        if new_status == 'AVAILABLE':
                            auto_returned_count = 0
                            active_rentals = Rental.objects.filter(
                                equipment=equipment,
                                status__in=['RENTED', 'OVERDUE']
                            )
                            for rental in active_rentals:
                                rental.status = 'RETURNED'
                                rental.return_date = timezone.now()
                                rental.returned_to = request.user if request.user.is_staff else None
                                rental.save()
                                auto_returned_count += 1
                                # 반납 이력 기록
                                EquipmentHistory.objects.create(
                                    equipment=equipment,
                                    action='RETURNED',
                                    user=request.user,
                                    old_value={
                                        'rental_id': rental.id,
                                        'user_id': rental.user.id,
                                        'username': rental.user.username,
                                        'status': 'RENTED'
                                    },
                                    new_value={
                                        'status': 'AVAILABLE',
                                        'return_date': rental.return_date.isoformat(),
                                        'returned_to': request.user.username if request.user.is_staff else None
                                    },
                                    details=f"장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 반납 from {rental.user.username}"
                                )

                        # 상태 변경 저장
                        equipment.status = new_status
                        equipment.save()
                        
                        # 대여중 상태일 때 Rental 레코드 생성
                        if new_status == 'RENTED' and user:
                            from .models import Rental
                            rental = Rental.objects.create(
                                equipment=equipment,
                                user=user,
                                approved_by=request.user,
                                rental_date=timezone.now(),
                                due_date=timezone.now() + timezone.timedelta(days=30),  # 기본 30일
                                status='RENTED',
                                notes=reason or f"일괄 상태 변경으로 대여"
                            )
                            logger.info(f"대여 레코드 생성: 장비 {equipment.asset_number} → 사용자 {user.username}")
                        
                        # 이력 기록
                        old_status_display = status_display_map.get(old_status, old_status)
                        details = f"장비 상태 변경: {old_status_display} → {new_status_display}"
                        if new_status == 'AVAILABLE':
                            # 자동 반납 건 수 표시
                            auto_count = locals().get('auto_returned_count', 0)
                            if auto_count and auto_count > 0:
                                details += f" (자동 반납 {auto_count}건)"
                        if new_status == 'RENTED' and user:
                            details += f" (대여자: {user.username})"
                        if reason:
                            details += f" (사유: {reason})"
                        
                        EquipmentHistory.objects.create(
                            equipment=equipment,
                            action='STATUS_CHANGED',
                            user=request.user,
                            old_value={'status': old_status},
                            new_value={
                                'status': new_status, 
                                'reason': reason,
                                'user_id': user.id if user else None,
                                'username': user.username if user else None
                            },
                            details=details
                        )
                        
                        success_count += 1
                        logger.info(f"장비 상태 변경 성공: {equipment.asset_number} {old_status} → {new_status}")
                        
                    except Exception as e:
                        error_msg = f"장비 {equipment.asset_number or equipment.model_name}: {str(e)}"
                        errors.append(error_msg)
                        failed_count += 1
                        logger.error(f"장비 상태 변경 실패: {equipment.asset_number} - {e}")
            
            # 결과 반환
            result = {
                "success": True,
                "message": f"{success_count}개 장비의 상태가 {new_status_display}로 변경되었습니다.",
                "success_count": success_count,
                "failed_count": failed_count,
                "total_count": len(equipment_ids),
                "new_status": new_status,
                "new_status_display": new_status_display
            }
            
            if errors:
                result["errors"] = errors
            
            if failed_count > 0:
                result["message"] += f" ({failed_count}개 실패)"
            
            return Response(result)
            
        except Exception as e:
            logger.error(f"일괄 상태 변경 중 오류: {e}")
            return Response({"detail": f"일괄 상태 변경 중 오류가 발생했습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """장비목록 엑셀 출력 (관리자 전용)"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        logger = logging.getLogger(__name__)
        try:
            logger.info("엑셀 내보내기 시작")
            
            # 모든 장비 정보 조회 (관리번호 오름차순 정렬)
            equipment_list = Equipment.objects.all().prefetch_related('rentals').order_by('management_number')
            logger.info(f"조회된 장비 수: {equipment_list.count()}")
            
            # 데이터 준비
            data = []
            for idx, equipment in enumerate(equipment_list, 1):
                try:
                    # 현재 대여 중인 정보 찾기
                    current_rental = equipment.rentals.filter(status='RENTED').first()
                    logger.debug(f"장비 {equipment.asset_number}의 현재 대여 정보: {current_rental}")
                    
                    # 사용자 이름 생성
                    user_name = ''
                    if current_rental and current_rental.user:
                        first_name = current_rental.user.first_name or ''
                        last_name = current_rental.user.last_name or ''
                        user_name = f"{last_name}{first_name}".strip() or current_rental.user.username
                    
                    equipment_data = {
                        '연번': idx,
                        '물품번호': equipment.asset_number or '',
                        '일련번호': equipment.serial_number or '',
                        '종류': equipment.get_equipment_type_display(),
                        '관리번호': equipment.management_number or '',
                        '제조사': equipment.manufacturer or '',
                        '모델명': equipment.model_name or '',
                        '구매일자': equipment.purchase_date.strftime('%Y-%m-%d') if equipment.purchase_date else '',
                        '구매금액': f"₩{int(equipment.purchase_price):,}" if equipment.purchase_price and equipment.purchase_price > 0 else '',
                        '대여자': user_name,
                        '대여자 아이디': current_rental.user.username if current_rental and current_rental.user else '',
                        '대여일자': current_rental.rental_date.strftime('%Y-%m-%d') if current_rental and current_rental.rental_date else '',
                        '관리자 확인': ''
                    }
                    data.append(equipment_data)
                    logger.debug(f"장비 {equipment.asset_number} 데이터 처리 완료")
                except Exception as e:
                    logger.error(f"장비 {equipment.asset_number} 데이터 처리 중 오류: {str(e)}")
                    # 오류가 발생해도 계속 진행하도록 수정
                    continue
            
            logger.info("DataFrame 생성 시작")
            # DataFrame 생성
            df = pd.DataFrame(data)
            logger.info(f"DataFrame 생성 완료: {len(df)} 행")
            
            # 엑셀 파일 생성
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=equipment_list_{datetime.now().strftime("%Y%m%d")}.xlsx'
            
            logger.info("엑셀 파일 저장 시작")
            # 엑셀 파일로 저장
            df.to_excel(response, index=False, sheet_name='장비목록')
            logger.info("엑셀 파일 저장 완료")
            
            return response
            
        except Exception as e:
            logger.error(f"엑셀 내보내기 중 오류 발생: {str(e)}", exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='bulk-update-excel')
    def bulk_update_excel(self, request):
        """엑셀 파일을 통한 장비 정보 일괄 등록/업데이트 (관리자 전용)"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        logger = logging.getLogger(__name__)
        
        try:
            # 파일이 업로드되었는지 확인
            if 'file' not in request.FILES:
                return Response(
                    {'error': '엑셀 파일이 업로드되지 않았습니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            uploaded_file = request.FILES['file']
            
            # 파일 확장자 확인
            if not uploaded_file.name.endswith(('.xlsx', '.xls')):
                return Response(
                    {'error': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            logger.info(f"엑셀 파일 업로드 시작: {uploaded_file.name}")
            
            # 엑셀 파일 읽기
            try:
                df = pd.read_excel(uploaded_file, sheet_name=0)
                logger.info(f"엑셀 파일 읽기 완료: {len(df)} 행")
            except Exception as e:
                logger.error(f"엑셀 파일 읽기 실패: {str(e)}")
                return Response(
                    {'error': f'엑셀 파일을 읽을 수 없습니다: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 필수 컬럼 확인
            required_columns = ['일련번호']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'필수 컬럼이 누락되었습니다: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 처리 가능한 컬럼 정의
            processable_columns = {
                '물품번호': 'asset_number',
                '일련번호': 'serial_number',
                '종류': 'equipment_type',
                '관리번호': 'management_number',
                '제조사': 'manufacturer',
                '모델명': 'model_name',
                '구매일자': 'purchase_date',
                '구매금액': 'purchase_price',
                '설명': 'description',
                '상태': 'status',
                '취득일': 'acquisition_date',
                '생산년도': 'manufacture_year'
            }
            
            # 장비 유형 매핑
            equipment_type_mapping = {
                '노트북': 'LAPTOP',
                '맥북': 'MACBOOK',
                '태블릿': 'TABLET',
                '데스크톱': 'DESKTOP',
                '모니터': 'MONITOR',
                '프린터': 'PRINTER',
                '프로젝터': 'PROJECTOR',
                '기타': 'OTHER'
            }
            
            # 상태 매핑
            status_mapping = {
                '사용 가능': 'AVAILABLE',
                '대여 중': 'RENTED',
                '점검 중': 'MAINTENANCE',
                '고장': 'BROKEN',
                '분실': 'LOST',
                '폐기': 'DISPOSED'
            }
            
            # 결과 통계
            results = {
                'total_rows': len(df),
                'created': 0,
                'updated': 0,
                'errors': [],
                'created_equipment': [],
                'updated_equipment': []
            }
            
            # 각 행 처리
            for index, row in df.iterrows():
                try:
                    serial_number = str(row['일련번호']).strip()
                    if not serial_number or serial_number == 'nan':
                        results['errors'].append(f"행 {index + 2}: 일련번호가 비어있습니다.")
                        continue
                    
                    # 일련번호로 장비 찾기
                    try:
                        equipment = Equipment.objects.get(serial_number=serial_number)
                        is_update = True
                    except Equipment.DoesNotExist:
                        equipment = None
                        is_update = False
                    
                    # 처리할 필드들 준비
                    processed_fields = {}
                    processed_field_names = []
                    
                    for excel_col, model_field in processable_columns.items():
                        if excel_col in df.columns and pd.notna(row[excel_col]):
                            value = row[excel_col]
                            
                            # 특별 처리
                            if excel_col == '종류':
                                # 한국어 -> 영어 코드 변환
                                if value in equipment_type_mapping:
                                    value = equipment_type_mapping[value]
                                elif value not in [choice[0] for choice in Equipment.EQUIPMENT_TYPE_CHOICES]:
                                    results['errors'].append(f"행 {index + 2}: 잘못된 장비 유형 '{row[excel_col]}'")
                                    continue
                            
                            elif excel_col == '상태':
                                # 한국어 -> 영어 코드 변환
                                if value in status_mapping:
                                    value = status_mapping[value]
                                elif value not in [choice[0] for choice in Equipment.STATUS_CHOICES]:
                                    results['errors'].append(f"행 {index + 2}: 잘못된 상태 '{row[excel_col]}'")
                                    continue
                            
                            elif excel_col in ['구매일자', '취득일']:
                                # 날짜 형식 처리
                                if isinstance(value, str):
                                    try:
                                        value = pd.to_datetime(value).date()
                                    except:
                                        results['errors'].append(f"행 {index + 2}: 잘못된 날짜 형식 '{value}'")
                                        continue
                                elif hasattr(value, 'date'):
                                    value = value.date()
                            
                            elif excel_col == '구매금액':
                                # 금액 처리 (₩ 기호 제거, 쉼표 제거)
                                if isinstance(value, str):
                                    value = value.replace('₩', '').replace(',', '').replace(' ', '')
                                    try:
                                        value = float(value)
                                    except:
                                        results['errors'].append(f"행 {index + 2}: 잘못된 금액 형식 '{row[excel_col]}'")
                                        continue
                            
                            elif excel_col == '생산년도':
                                # 정수 처리
                                try:
                                    value = int(value)
                                except:
                                    results['errors'].append(f"행 {index + 2}: 잘못된 생산년도 형식 '{value}'")
                                    continue
                            
                            processed_fields[model_field] = value
                            processed_field_names.append(excel_col)
                    
                    if is_update:
                        # 기존 장비 업데이트
                        if processed_fields:
                            for field, value in processed_fields.items():
                                setattr(equipment, field, value)
                            equipment.save()
                            results['updated'] += 1
                            results['updated_equipment'].append({
                                'serial_number': equipment.serial_number,
                                'asset_number': equipment.asset_number,
                                'updated_fields': processed_field_names
                            })
                            logger.info(f"장비 업데이트 완료: {equipment.serial_number} - {', '.join(processed_field_names)}")
                    else:
                        # 새 장비 생성
                        # 필수 필드 확인
                        if 'equipment_type' not in processed_fields:
                            results['errors'].append(f"행 {index + 2}: 장비 유형이 필요합니다.")
                            continue
                        
                        # 기본값 설정
                        equipment_data = {
                            'serial_number': serial_number,
                            'equipment_type': processed_fields.get('equipment_type', 'OTHER'),
                            'status': processed_fields.get('status', 'AVAILABLE'),
                            'description': processed_fields.get('description', ''),
                            'acquisition_date': processed_fields.get('acquisition_date', timezone.now().date())
                        }
                        
                        # 나머지 필드 추가
                        for field, value in processed_fields.items():
                            if field not in equipment_data:
                                equipment_data[field] = value
                        
                        # 장비 생성
                        serializer = EquipmentSerializer(data=equipment_data)
                        if serializer.is_valid():
                            serializer.save()
                            results['created'] += 1
                            results['created_equipment'].append({
                                'serial_number': equipment_data['serial_number'],
                                'asset_number': equipment_data.get('asset_number', ''),
                                'created_fields': processed_field_names
                            })
                            logger.info(f"장비 생성 완료: {equipment_data['serial_number']}")
                        else:
                            results['errors'].append(f"행 {index + 2}: 장비 생성 실패 - {serializer.errors}")
                
                except Exception as e:
                    logger.error(f"행 {index + 2} 처리 중 오류: {str(e)}")
                    results['errors'].append(f"행 {index + 2}: {str(e)}")
            
            logger.info(f"일괄 처리 완료: 총 {results['total_rows']}행, 생성 {results['created']}개, 업데이트 {results['updated']}개")
            
            return Response({
                'success': True,
                'message': f'일괄 처리가 완료되었습니다. (생성: {results["created"]}개, 업데이트: {results["updated"]}개)',
                'results': results
            })
            
        except Exception as e:
            logger.error(f"일괄 처리 중 오류 발생: {str(e)}", exc_info=True)
            return Response(
                {'error': f'일괄 처리 중 오류가 발생했습니다: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EquipmentReadOnlyViewSet(viewsets.ReadOnlyModelViewSet):
    """
    장비 조회 API (사용자용 - 읽기 전용)
    """
    queryset = Equipment.objects.all()
    serializer_class = EquipmentLiteSerializer
    # permission_classes 제거 - 기본 권한 클래스 사용
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        'asset_number', 
        'serial_number', 
        'description', 
        'manufacturer', 
        'model_name', 
        'management_number'
    ]
    ordering_fields = ['asset_number', 'equipment_type', 'status', 'acquisition_date']
    pagination_class = EquipmentPagination
    
    def get_queryset(self):
        queryset = Equipment.objects.prefetch_related(
            Prefetch(
                'rentals',
                queryset=Rental.objects.filter(status='RENTED').select_related('user').order_by('-rental_date'),
                to_attr='current_rentals'
            )
        )
        
        # 상태 필터링
        status = self.request.query_params.get('status', None)
        if status:
            queryset = queryset.filter(status=status)
        
        # 장비 타입 필터링
        equipment_type = self.request.query_params.get('equipment_type', None)
        if equipment_type:
            queryset = queryset.filter(equipment_type=equipment_type)
        
        return queryset
    
    # get_permissions 메서드 제거 - 중앙화된 권한 관리 사용
    
    @action(detail=False, methods=['get'])
    def available(self, request):
        """대여 가능한 장비 목록 조회 (인증된 사용자 접근 가능)"""
        # 대여 가능한 장비만 필터링하되 대여 정보도 함께 로드
        available_equipment = Equipment.objects.filter(status='AVAILABLE').prefetch_related(
            Prefetch(
                'rentals',
                queryset=Rental.objects.filter(status='RENTED').select_related('user').order_by('-rental_date'),
                to_attr='current_rentals'
            )
        ).order_by('management_number')
        
        # 페이지네이션 없이 배열로 반환
        serializer = self.get_serializer(available_equipment, many=True)
        return Response(serializer.data)



    @action(detail=False, methods=['get', 'post'], url_path='by-mac')
    def by_mac(self, request):
        """MAC 주소로 장비를 조회하는 API"""
        # GET 요청 처리 (쿼리 파라미터에서 MAC 주소 추출)
        if request.method == 'GET':
            mac_address = request.query_params.get('mac_address')
            if not mac_address:
                return Response(
                    {'error': 'mac_address 파라미터가 필요합니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            mac_addresses = [mac_address]
        
        # POST 요청 처리 (요청 본문에서 MAC 주소 목록 추출)
        else:
            mac_address = request.data.get('mac_address')
            mac_addresses = request.data.get('mac_addresses', [])
            
            if not mac_address and not mac_addresses:
                return Response(
                    {'error': 'mac_address 또는 mac_addresses 배열이 필요합니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            # 단일 MAC 주소가 있으면 목록에 추가
            if mac_address:
                mac_addresses.append({'mac_address': mac_address})
            elif isinstance(mac_addresses, list) and all(isinstance(item, dict) for item in mac_addresses):
                # mac_addresses가 객체 배열인 경우 MAC 주소만 추출
                mac_addresses = [item.get('mac_address') for item in mac_addresses if item.get('mac_address')]
        
        # MAC 주소 목록 순회하며 장비 찾기
        found_equipment = []
        for current_mac in mac_addresses:
            try:
                mac_obj = EquipmentMacAddress.objects.get(mac_address=current_mac)
                equipment = mac_obj.equipment
                if equipment not in found_equipment:  # 중복 제거
                    found_equipment.append(equipment)
            except EquipmentMacAddress.DoesNotExist:
                continue
        
        if not found_equipment:
            return Response(
                {'error': '입력한 MAC 주소를 가진 장비를 찾을 수 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # 결과 반환
        serializer = self.get_serializer(found_equipment, many=True)
        return Response(serializer.data)


class RentalViewSet(viewsets.ModelViewSet):
    """
    대여 정보 관리 API
    """
    queryset = Rental.objects.all()
    serializer_class = RentalSerializer
    # permission_classes 제거 - 기본 권한 클래스 사용
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['equipment__asset_number', 'equipment__serial_number', 'user__username']
    ordering_fields = ['rental_date', 'due_date', 'return_date', 'status']
    http_method_names = ['get', 'post', 'put', 'patch', 'delete']
    
    def get_permissions(self):
        """
        액션별로 권한 설정:
        - 관리자는 모든 권한
        - 일반 사용자는 자신의 대여 정보만 접근 가능
        """
        logger = logging.getLogger('rentals')
        logger.debug(f"get_permissions 호출됨: action={self.action}, method={self.request.method}")
        
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            logger.debug("관리자 권한 필요")
            return [permissions.IsAdminUser()]
        logger.debug("인증된 사용자 권한")
        return [permissions.IsAuthenticated()]
    
    def get_queryset(self):
        user = self.request.user
        
        if user.is_staff:
            return Rental.objects.all()
        return Rental.objects.filter(user=user)
    
    def create(self, request, *args, **kwargs):
        """대여 정보 생성"""
        logger = logging.getLogger('rentals')
        logger.debug(f"create 메서드 호출됨: {request.data}")
        
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            logger.error(f"유효성 검사 오류: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                # 장비 상태 체크
                equipment = serializer.validated_data.get('equipment')
                if equipment.status != 'AVAILABLE':
                    logger.warning(f"대여 불가능한 장비: {equipment.id}, 상태: {equipment.status}")
                    return Response(
                        {"detail": f"대여할 수 없는 장비입니다. 현재 상태: {equipment.get_status_display()}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # 공통 서비스 사용
                from .services.rental_logic import create_rental
                user = serializer.validated_data.get('user', request.user)
                rental = create_rental(
                    equipment=equipment,
                    user=user,
                    approved_by=request.user,
                    due_date=serializer.validated_data.get('due_date'),
                    notes=serializer.validated_data.get('notes', ''),
                )
                
                logger.info(f"대여 생성 성공: {rental.id}")
                response_serializer = self.get_serializer(rental)
                return Response({
                    "success": True,
                    "data": response_serializer.data,
                    "message": "대여가 성공적으로 생성되었습니다."
                }, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            logger.error(f"대여 생성 중 오류: {e}")
            return Response(
                {"detail": f"대여 생성 중 오류가 발생했습니다: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    

    
    @action(detail=False, methods=['get'], url_path='active')
    def my_rentals(self, request):
        """내 대여 목록 조회 (현재 대여 중인 것만)"""
        my_rentals = Rental.objects.filter(
            user=request.user, 
            status='RENTED'
        ).select_related('equipment', 'user').prefetch_related(
            'equipment__rental_requests'
        ).order_by('-rental_date')
        page = self.paginate_queryset(my_rentals)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(my_rentals, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='history')
    def my_history(self, request):
        """내 대여 이력 전체 조회 (과거 대여 포함)"""
        my_history = Rental.objects.filter(user=request.user).select_related('equipment', 'user').order_by('-rental_date')
        page = self.paginate_queryset(my_history)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(my_history, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def return_equipment(self, request, pk=None):
        """장비 반납 처리"""
        rental = self.get_object()
        
        if rental.status not in ['RENTED', 'OVERDUE']:
            return Response(
                {"detail": "이미 반납된 장비입니다."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                # 공통 서비스 사용
                from .services.rental_logic import return_rental
                return_rental(rental, request.user)
                
                serializer = self.get_serializer(rental)
                return Response({
                    "success": True,
                    "data": serializer.data,
                    "message": "장비가 성공적으로 반납되었습니다."
                })
        except Exception as e:
            return Response(
                {"detail": f"반납 처리 중 오류가 발생했습니다: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class RentalRequestViewSet(viewsets.ModelViewSet):
    """
    대여/반납 요청 관리 API
    """
    queryset = RentalRequest.objects.all()
    serializer_class = RentalRequestSerializer
    # permission_classes 제거 - 기본 권한 클래스 사용
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['equipment__asset_number', 'user__username', 'reason']
    ordering_fields = ['requested_date', 'status', 'request_type']
    
    def get_queryset(self):
        user = self.request.user
        
        if user.is_staff:
            return RentalRequest.objects.all()
        return RentalRequest.objects.filter(user=user)
    
    def create(self, request, *args, **kwargs):
        """대여/반납 요청 생성 - equipment 필드 처리"""
        logger = logging.getLogger('rentals')
        
        logger.debug(f"대여/반납 요청 API 호출됨: {request.data}")
        data = request.data.copy()
        
        # 요청 데이터 로깅
        logger.debug(f"원본 요청 데이터: {request.data}")
        
        # 중복 요청 체크
        equipment_id = data.get('equipment')
        request_type = data.get('request_type')
        
        if equipment_id and request_type:
            existing_request = RentalRequest.objects.filter(
                equipment_id=equipment_id,
                request_type=request_type,
                status='PENDING'
            ).first()
            
            if existing_request:
                return Response({
                    "detail": "이미 진행 중인 요청이 있습니다.",
                    "existing_request": {
                        "id": existing_request.id,
                        "request_type": existing_request.get_request_type_display(),
                        "status": existing_request.get_status_display(),
                        "requested_date": existing_request.requested_date
                    }
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # 이 부분을 수정하지 않음 - equipment 필드를 그대로 유지
        # 시리얼라이저는 equipment 필드를 사용하므로 변환하지 않음
        
        logger.debug(f"처리할 데이터: {data}")
        
        serializer = self.get_serializer(data=data)
        
        # 유효성 검사 수행
        if not serializer.is_valid():
            logger.error(f"유효성 검사 오류: {serializer.errors}")
            return Response({"errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response({"success": True, "data": serializer.data}, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            logger.exception(f"대여/반납 요청 생성 중 오류 발생: {str(e)}")
            return Response({"success": False, "error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def perform_create(self, serializer):
        """요청 생성 시 현재 사용자 정보 추가"""
        serializer.save(user=self.request.user)
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """승인 대기 중인 요청 목록 조회 (관리자용)"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        pending_requests = RentalRequest.objects.filter(status='PENDING')
        serializer = self.get_serializer(pending_requests, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """요청 승인 처리"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        rental_request = self.get_object()
        
        if rental_request.status != 'PENDING':
            return Response(
                {"detail": "이미 처리된 요청입니다."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 요청 정보 업데이트
        rental_request.status = 'APPROVED'
        rental_request.processed_by = request.user
        rental_request.processed_at = timezone.now()
        
        if rental_request.request_type == 'RENT':
            # 장비 상태 체크
            equipment = rental_request.equipment
            if equipment.status != 'AVAILABLE':
                return Response(
                    {"detail": f"대여할 수 없는 장비입니다. 현재 상태: {equipment.get_status_display()}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            # 중복된 대여 기록 정리
            duplicate_rentals = Rental.objects.filter(
                user=rental_request.user,
                equipment=equipment,
                status__in=['RENTED', 'OVERDUE']
            ).order_by('-rental_date')
            
            if duplicate_rentals.exists():
                # 가장 최근 기록을 제외한 나머지 삭제
                latest_rental = duplicate_rentals.first()
                duplicate_rentals.exclude(id=latest_rental.id).delete()
                logger.info(f"중복된 대여 기록 정리 완료: 장비={equipment.id}, 사용자={rental_request.user.id}")
                
                # 기존 대여 기록 반납 처리
                latest_rental.status = 'RETURNED'
                latest_rental.return_date = timezone.now()
                latest_rental.returned_to = request.user
                latest_rental.save()
            
            # 새 대여 정보 생성
            rental = Rental.objects.create(
                user=rental_request.user,
                equipment=equipment,
                rental_date=timezone.now(),
                due_date=rental_request.expected_return_date or (timezone.now() + timezone.timedelta(days=30)),
                status='RENTED',
                notes=f"대여 요청 #{rental_request.id}에 의해 생성됨",
                approved_by=request.user
            )
            
            # 요청과 대여 정보 연결
            rental_request.rental = rental
            
            # 장비 상태 업데이트
            equipment.status = 'RENTED'
            equipment.save()
            
            # 장비 대여 이력 기록
            EquipmentHistory.objects.create(
                equipment=equipment,
                action='RENTED',
                user=request.user,
                new_value={
                    'rental_id': rental.id,
                    'user_id': rental.user.id,
                    'username': rental.user.username,
                    'rental_date': rental.rental_date.isoformat(),
                    'due_date': rental.due_date.isoformat(),
                    'status': 'RENTED'
                },
                details=f"장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 대여 승인 to {rental.user.username}"
            )
            
        elif rental_request.request_type == 'RETURN':
            # 현재 사용자의 해당 장비 대여 정보 검색
            try:
                rental = Rental.objects.get(
                    user=rental_request.user,
                    equipment=rental_request.equipment,
                    status__in=['RENTED', 'OVERDUE']
                )
            except Rental.DoesNotExist:
                return Response(
                    {"detail": "대여 중인 장비를 찾을 수 없습니다."},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            # 대여 정보 업데이트
            rental.status = 'RETURNED'
            rental.return_date = timezone.now()
            rental.returned_to = request.user
            rental.save()
            
            # 요청과 대여 정보 연결
            rental_request.rental = rental
            
            # 장비 상태 업데이트
            equipment = rental_request.equipment
            equipment.status = 'AVAILABLE'
            equipment.save()
            
            # 장비 반납 이력 기록
            EquipmentHistory.objects.create(
                equipment=equipment,
                action='RETURNED',
                user=request.user,
                old_value={
                    'rental_id': rental.id,
                    'user_id': rental.user.id,
                    'username': rental.user.username,
                    'status': 'RENTED'
                },
                new_value={
                    'status': 'AVAILABLE',
                    'return_date': rental.return_date.isoformat(),
                    'returned_to': request.user.username if request.user.is_staff else None
                },
                details=f"장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 반납 from {rental.user.username}"
            )
        
        rental_request.save()
        serializer = self.get_serializer(rental_request)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """요청 거부 처리"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        rental_request = self.get_object()
        
        if rental_request.status != 'PENDING':
            return Response(
                {"detail": "이미 처리된 요청입니다."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        rental_request.status = 'REJECTED'
        rental_request.processed_by = request.user
        rental_request.processed_at = timezone.now()
        
        # 거부 사유가 있으면 저장
        reason = request.data.get('reason')
        if reason:
            rental_request.reject_reason = reason
        
        rental_request.save()
        
        serializer = self.get_serializer(rental_request)
        return Response({
            "success": True,
            "data": serializer.data,
            "message": f"요청이 성공적으로 {request.data.get('action')}되었습니다."
        })
        
    @action(detail=True, methods=['post'])
    def process(self, request, pk=None):
        """요청 처리 (승인 또는 거부)"""
        logger = logging.getLogger('rentals')
        logger.debug(f"process 액션 호출됨: {request.data}")
        
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        rental_request = self.get_object()
        
        if rental_request.status != 'PENDING':
            return Response(
                {"detail": "이미 처리된 요청입니다."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 액션 타입 확인 (approve 또는 reject)
        action = request.data.get('action')
        reason = request.data.get('reason')
        
        logger.debug(f"처리 유형: {action}, 사유: {reason}")
        
        if action not in ['approve', 'reject']:
            return Response(
                {"detail": "유효하지 않은 액션입니다. 'approve' 또는 'reject'만 허용됩니다."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 승인 처리
        if action == 'approve':
            rental_request.status = 'APPROVED'
            rental_request.processed_by = request.user
            rental_request.processed_at = timezone.now()
            
            if rental_request.request_type == 'RENTAL':
                # 장비 상태 체크
                equipment = rental_request.equipment
                if equipment.status != 'AVAILABLE':
                    return Response(
                        {"detail": f"대여할 수 없는 장비입니다. 현재 상태: {equipment.get_status_display()}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                    
                # 중복된 대여 기록 정리
                duplicate_rentals = Rental.objects.filter(
                    user=rental_request.user,
                    equipment=equipment,
                    status__in=['RENTED', 'OVERDUE']
                ).order_by('-rental_date')
                
                if duplicate_rentals.exists():
                    # 가장 최근 기록을 제외한 나머지 삭제
                    latest_rental = duplicate_rentals.first()
                    duplicate_rentals.exclude(id=latest_rental.id).delete()
                    logger.info(f"중복된 대여 기록 정리 완료: 장비={equipment.id}, 사용자={rental_request.user.id}")
                
                # 새 대여 정보 생성
                rental = Rental.objects.create(
                    user=rental_request.user,
                    equipment=equipment,
                    rental_date=timezone.now(),
                    due_date=rental_request.expected_return_date or (timezone.now() + timezone.timedelta(days=30)),
                    status='RENTED',
                    notes=f"대여 요청 #{rental_request.id}에 의해 생성됨",
                    approved_by=request.user
                )
                
                # 요청과 대여 정보 연결
                rental_request.rental = rental
                
                # 장비 상태 업데이트
                equipment.status = 'RENTED'
                equipment.save()
                
                # 장비 대여 이력 기록
                EquipmentHistory.objects.create(
                    equipment=equipment,
                    action='RENTED',
                    user=request.user,
                    new_value={
                        'rental_id': rental.id,
                        'user_id': rental.user.id,
                        'username': rental.user.username,
                        'rental_date': rental.rental_date.isoformat(),
                        'due_date': rental.due_date.isoformat(),
                        'status': 'RENTED'
                    },
                    details=f"장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 대여 승인 to {rental.user.username}"
                )
                
            elif rental_request.request_type == 'RETURN':
                # 중복된 대여 기록 정리 후 가장 최근 대여 정보 검색
                rentals = Rental.objects.filter(
                    user=rental_request.user,
                    equipment=rental_request.equipment,
                    status__in=['RENTED', 'OVERDUE']
                ).order_by('-rental_date')
                
                if not rentals.exists():
                    return Response(
                        {"detail": "대여 중인 장비를 찾을 수 없습니다."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # 가장 최근 대여 기록만 남기고 나머지 삭제
                rental = rentals.first()
                rentals.exclude(id=rental.id).delete()
                logger.info(f"중복된 대여 기록 정리 완료: 장비={rental_request.equipment.id}, 사용자={rental_request.user.id}")
                    
                # 대여 정보 업데이트
                rental.status = 'RETURNED'
                rental.return_date = timezone.now()
                rental.returned_to = request.user
                rental.save()
                
                # 요청과 대여 정보 연결
                rental_request.rental = rental
                
                # 장비 상태 업데이트
                equipment = rental_request.equipment
                equipment.status = 'AVAILABLE'
                equipment.save()
                
                # 장비 반납 이력 기록
                EquipmentHistory.objects.create(
                    equipment=equipment,
                    action='RETURNED',
                    user=request.user,
                    old_value={
                        'rental_id': rental.id,
                        'user_id': rental.user.id,
                        'username': rental.user.username,
                        'status': 'RENTED'
                    },
                    new_value={
                        'status': 'AVAILABLE',
                        'return_date': rental.return_date.isoformat(),
                        'returned_to': request.user.username if request.user.is_staff else None
                    },
                    details=f"장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 반납 from {rental.user.username}"
                )
            
            rental_request.save()
            serializer = self.get_serializer(rental_request)
            return Response({
                "success": True,
                "data": serializer.data,
                "message": f"요청이 성공적으로 {action}되었습니다."
            })
        
        # 거부 처리
        elif action == 'reject':
            rental_request.status = 'REJECTED'
            rental_request.processed_by = request.user
            rental_request.processed_at = timezone.now()
            
            # 거부 사유가 있으면 저장
            if reason:
                rental_request.reject_reason = reason
            
            rental_request.save()
            
            serializer = self.get_serializer(rental_request)
            return Response({
                "success": True,
                "data": serializer.data,
                "message": f"요청이 성공적으로 {action}되었습니다."
            })
    
    @action(detail=False, methods=['get'], url_path='my')
    def my_requests(self, request):
        """내 대여/반납 요청 목록 조회"""
        my_requests = RentalRequest.objects.filter(user=request.user).order_by('-requested_date')
        serializer = self.get_serializer(my_requests, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['patch'])
    def update_reason(self, request, pk=None):
        """요청 사유 수정 (PENDING 상태인 경우만 가능)"""
        logger = logging.getLogger('rentals')
        logger.debug(f"update_reason 호출됨: {request.data}")
        
        rental_request = self.get_object()
        
        # 요청자 본인 또는 관리자만 수정 가능
        if rental_request.user != request.user and not request.user.is_staff:
            return Response({"detail": "이 요청을 수정할 권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        
        # 대기 상태일 때만 수정 가능
        if rental_request.status != 'PENDING':
            return Response({"detail": "이미 처리된 요청은 수정할 수 없습니다."}, status=status.HTTP_400_BAD_REQUEST)
        
        # 사유 업데이트
        reason = request.data.get('reason')
        if reason is not None:
            rental_request.reason = reason
            rental_request.save(update_fields=['reason'])
            
            logger.debug(f"사유 업데이트 성공: {reason}")
            
            serializer = self.get_serializer(rental_request)
            return Response({
                "success": True,
                "data": serializer.data,
                "message": "요청 사유가 성공적으로 업데이트되었습니다."
            })
        else:
            return Response({"detail": "사유가 제공되지 않았습니다."}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def resubmit(self, request, pk=None):
        """거부된 요청 재신청 (REJECTED 상태인 경우만 가능)"""
        logger = logging.getLogger('rentals')
        logger.debug(f"resubmit 호출됨: {request.data}")
        
        rental_request = self.get_object()
        
        # 요청자 본인만 재신청 가능
        if rental_request.user != request.user:
            return Response({"detail": "이 요청을 재신청할 권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        
        # 거부된 상태일 때만 재신청 가능
        if rental_request.status != 'REJECTED':
            return Response({"detail": "거부된 요청만 재신청할 수 있습니다."}, status=status.HTTP_400_BAD_REQUEST)
        
        # 새 사유 업데이트 (제공된 경우)
        reason = request.data.get('reason')
        if reason:
            rental_request.reason = reason
        
        # 상태 업데이트 및 처리 정보 초기화
        rental_request.status = 'PENDING'
        rental_request.processed_by = None
        rental_request.processed_at = None
        rental_request.requested_date = timezone.now()  # 요청일을 현재 시간으로 업데이트
        
        # 반납 예정일 업데이트 (제공된 경우)
        expected_return_date = request.data.get('expected_return_date')
        if expected_return_date and rental_request.request_type == 'RENT':
            try:
                rental_request.expected_return_date = expected_return_date
            except:
                pass
        
        rental_request.save()
        
        logger.debug(f"재신청 성공: {rental_request.id}")
        
        serializer = self.get_serializer(rental_request)
        return Response({
            "success": True,
            "data": serializer.data,
            "message": "요청이 성공적으로 재신청되었습니다."
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """요청 목록 엑셀 출력"""
        if not request.user.is_staff:
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            
        request_list = RentalRequest.objects.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "대여/반납 요청 내역"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        
        # 컬럼 헤더 설정
        columns = ['ID', '사용자', '요청 유형', '장비명', '요청일', '상태', '처리자', '처리일']
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 입력
        for row_num, req in enumerate(request_list, 2):
            ws.cell(row=row_num, column=1, value=req.id)
            ws.cell(row=row_num, column=2, value=req.user.username)
            ws.cell(row=row_num, column=3, value=req.get_request_type_display())
            ws.cell(row=row_num, column=4, value=req.equipment.asset_number)
            ws.cell(row=row_num, column=5, value=str(req.requested_date))
            ws.cell(row=row_num, column=6, value=req.get_status_display())
            ws.cell(row=row_num, column=7, value=req.processed_by.username if req.processed_by else "미처리")
            ws.cell(row=row_num, column=8, value=str(req.processed_at) if req.processed_at else "미처리")
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="rental_requests.xlsx"'
        
        return response


class EquipmentMacAddressViewSet(viewsets.ModelViewSet):
    queryset = EquipmentMacAddress.objects.all()
    serializer_class = EquipmentMacAddressSerializer
    permission_classes = [AllowAny]  # MAC 주소 관련은 누구나 접근 가능
    parser_classes = [JSONParser, FormParser, MultiPartParser]  # 파서 클래스 추가
    
    
    @action(detail=False, methods=['post'])
    def login_and_rent(self, request):
        """MAC 주소로 로그인하고 장비를 대여하는 API"""
        # 단일 MAC 주소 또는 MAC 주소 목록 받기
        mac_address = request.data.get('mac_address')
        mac_addresses = request.data.get('mac_addresses', [])
        
        # 둘 다 없는 경우
        if not mac_address and not mac_addresses:
            return Response(
                {'error': 'mac_address 또는 mac_addresses 배열이 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 단일 MAC 주소가 있으면 목록에 추가
        if mac_address:
            mac_addresses.append({'mac_address': mac_address})
        
        # MAC 주소 목록 순회하며 사용자 찾기
        for mac_item in mac_addresses:
            current_mac = mac_item.get('mac_address') if isinstance(mac_item, dict) else mac_item
            if not current_mac:
                continue
                
            try:
                # MAC 주소로 장비 찾기
                mac_obj = EquipmentMacAddress.objects.get(mac_address=current_mac)
                equipment = mac_obj.equipment
                
                # 현재 대여 중인 사용자 찾기
                current_rental = equipment.rentals.filter(status='RENTED').first()
                if current_rental:
                    return Response({
                        'message': '이미 등록된 장비입니다.',
                        'equipment': EquipmentSerializer(equipment).data,
                        'current_user': {
                            'username': current_rental.user.username,
                            'full_name': current_rental.user.get_full_name()
                        }
                    }, status=status.HTTP_200_OK)
                
                # 사용자 정보 확인
                if hasattr(equipment, 'user') and equipment.user:
                    user = equipment.user
                    login(request, user)
                    
                    # 대여 정보 생성
                    rental = Rental.objects.create(
                        user=user,
                        equipment=equipment,
                        rental_date=timezone.now(),
                        due_date=timezone.now() + timezone.timedelta(days=30),  # 30일 후 반납 예정
                        status='RENTED',
                        notes=f'MAC 주소({current_mac})를 통한 자동 대여'
                    )
                    
                    # 장비 상태 업데이트
                    equipment.status = 'RENTED'
                    equipment.save()
                    
                    # 장비 대여 이력 기록
                    EquipmentHistory.objects.create(
                        equipment=equipment,
                        action='RENTED',
                        user=request.user,
                        new_value={
                            'rental_id': rental.id,
                            'user_id': rental.user.id,
                            'username': rental.user.username,
                            'rental_date': rental.rental_date.isoformat(),
                            'due_date': rental.due_date.isoformat(),
                            'status': 'RENTED'
                        },
                        details=f"장비 '{equipment.asset_number or equipment.model_name or equipment.serial_number}' 대여 승인 to {rental.user.username}"
                    )
                    
                    return Response({
                        'message': '로그인 및 장비 대여가 완료되었습니다.',
                        'user': {
                            'id': user.id,
                            'username': user.username,
                            'first_name': user.first_name,
                            'last_name': user.last_name,
                            'email': user.email
                        },
                        'rental': RentalSerializer(rental).data,
                        'equipment': EquipmentSerializer(equipment).data
                    })
            except EquipmentMacAddress.DoesNotExist:
                # 이 MAC 주소로 장비를 찾지 못했으므로 다음 MAC 주소 시도
                continue
        
        # 모든 MAC 주소를 시도해도 장비를 찾지 못한 경우
        return Response(
            {'error': '등록된 MAC 주소를 가진 장비가 없습니다.'},
            status=status.HTTP_404_NOT_FOUND
        )
